"""
Build SS product import Excel from SS_CATALOG_MRP.pdf.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pdfplumber

ROOT = Path(__file__).resolve().parent
PDF_PATH = ROOT / "SS_CATALOG_MRP.pdf"
TEMPLATE_PATH = ROOT / "ss-product-import.xlsx"
CATEGORY_PATH = ROOT / "ss_category_import_template.xlsx"
OUTPUT_PATH = ROOT / "ss-product-import-filled.xlsx"

SKIP_NAMES = re.compile(
    r"^(voice of willow|best never stops|terms?\b|these terms|full(\s+no)?|"
    r"size\b|men|youth|boys|harrow|pack of|per pcs|pair\)?|xl\)?|red|blue|"
    r"white line|closure|tic|er\)|range|series)$",
    re.I,
)

# Explicit section header -> category
SECTION_MAP = [
    (r"english\s*willow|e\s*n\s*g\s*l\s*i\s*s\s*h\s*w\s*i\s*l\s*l\s*o\s*w|ew\s+junior|junior\s+english", "CC_EW_BATS", "English Willow Cricket Bat"),
    (r"kashmir\s*willow|kw\s+junior|junior\s*kw|master\s+junior\s+kw", "CC_KW_BATS", "Kashmir Willow Cricket Bat"),
    (r"plastic\s*(cr\.?\s*)?bats?|tennis\s*ball\s*bats?|composite\s*scoop|soft\s*pro|painted\s*bats?|scoop\s*bats?", "CC_TENNIS_BATS", "Tennis Ball Cricket Bat"),
    (r"cricket\s*helmets?|\bhelmets?\b", "CC_HELMET", "Cricket Helmet"),
    (r"super\s*caps?|\bcaps?\b", "CC_CAPS", "Cricket Cap"),
    (r"panama\s*hats?|\bhats?\b|bucket\s*hat", "CC_HATS", "Cricket Hat"),
    (r"wheelie\s*kitbags?", "CC_WHEELIE_KITBAGS", "Wheelie Cricket Kitbag"),
    (r"duffle\s*kitbags?|kitbags?", "CC_DUFFLE_KITBAGS", "Cricket Kitbag"),
    (r"cricket\s*kit", "CC_DUFFLE_KITBAGS", "Cricket Kit"),
    (r"leather\s*balls?", "CC_LEATHER_BALLS", "Leather Cricket Ball"),
    (r"tennis\s*balls?", "CC_TENNIS_BALLS", "Tennis Cricket Ball"),
    (r"synthetic\s*balls?", "CC_SYNTHETIC_BALLS", "Synthetic Cricket Ball"),
    (r"practice\s*balls?", "CC_PRACTICE_BALLS", "Practice Cricket Ball"),
    (r"^cricket\s*balls?$", "CC_LEATHER_BALLS", "Cricket Ball"),
    (r"\bstumps?\b|ball\s*mallet|rubber\s*base", "CC_STUMPS", "Cricket Stumps"),
    (r"\bshoes?\b", "CC_SHOES", "Cricket Shoes"),
    (r"\bsocks?\b", "CC_SOCKS", "Cricket Socks"),
    (r"combo\s*whites?", "CC_COMBO_WHITES", "Cricket Combo Whites"),
    (r"cricket\s*whites?|^whites?$", "CC_WHITES", "Cricket Whites"),
    (r"inner\s*gloves?|inner\s*skull", "CC_INNER_GLOVES", "Inner Cricket Gloves"),
    (r"batting\s*gloves?", "CC_BATTING_GLOVES", "Cricket Batting Gloves"),
    (r"keeping\s*gloves?|wicket\s*keeping\s*gloves?", "CC_KEEPING_GLOVES", "Wicket Keeping Gloves"),
    (r"batting\s*pads?|wicket\s*keeping\s*pads?|keeping\s*pads?", "CC_BATTING_PADS", "Cricket Pads"),
    (r"thigh\s*pads?|thigh\s*guards?", "CC_THIGH_PADS", "Cricket Thigh Pad"),
    (r"elbow\s*guards?|elbow\s*sleeves?", "CC_ELBOW_GUARDS", "Cricket Elbow Guard"),
    (r"supporters?|abdo\s*guard", "CC_SUPPORTERS", "Cricket Supporter"),
    (r"cricket\s*bats?|^core\s*range$|^ton\s*range$|^sky\s*(series|range)$|^thala\s*range$|^valarie\s*range$|^devil\s*range$|^gg\s*smacker|^master\s*series|^va-?900|^vintage$|^colt\s*range|^suryavanshi", "CC_EW_BATS", "Cricket Bat"),
]

PAGE_DEFAULTS = {
    2: ("CC_EW_BATS", "English Willow Cricket Bat"),
    3: ("CC_EW_BATS", "English Willow Cricket Bat"),
    4: ("CC_KW_BATS", "Kashmir Willow Cricket Bat"),
    5: ("CC_TENNIS_BATS", "Tennis Ball Cricket Bat"),
    6: ("CC_KW_BATS", "Kashmir Willow Cricket Bat"),
    7: ("CC_HELMET", "Cricket Helmet"),
    8: ("CC_WHEELIE_KITBAGS", "Wheelie Cricket Kitbag"),
    9: ("CC_LEATHER_BALLS", "Cricket Ball"),
    10: ("CC_STUMPS", "Cricket Stumps"),
    11: ("CC_SOCKS", "Cricket Socks"),
    12: ("CC_BATTING_GLOVES", "Cricket Batting Gloves"),
    13: ("CC_KEEPING_GLOVES", "Wicket Keeping Gloves"),
}


@dataclass
class ProductRow:
    name: str
    mrp: int
    category_id: str
    product_type_label: str
    size_label: str = ""
    section: str = ""
    source: str = "text"
    attrs: dict = field(default_factory=dict)


def clean(value: str) -> str:
    value = (value or "").replace("\u201c", '"').replace("\u201d", '"').replace("\u2019", "'")
    value = value.replace("�", '"')
    return re.sub(r"\s+", " ", value).strip(" -|/")


def parse_price(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if not s or s in {"-", "—", "–"}:
        return None
    if not re.fullmatch(r"\d{3,6}", s):
        return None
    return int(s)


def title_case(name: str) -> str:
    name = clean(name)
    keep = {
        "SS", "EW", "KW", "GG", "VA", "DK", "MS", "JR", "XL", "S/M/L",
        "NO.6", "NO.5", "NO.4", "NO.3", "PRO", "Ton",
    }
    words = []
    for w in name.split(" "):
        up = w.upper()
        if up in keep or up.replace("NO.", "NO.") in keep:
            words.append(up if up in {"SS", "EW", "KW", "GG", "VA", "DK", "MS", "JR", "XL", "PRO"} else w)
        elif re.fullmatch(r"\d+(\.\d+)?", w):
            words.append(w)
        elif "/" in w:
            words.append("/".join(title_case(p) if p else "" for p in w.split("/")))
        else:
            words.append(w[:1].upper() + w[1:].lower())
    return " ".join(words)


def slug_id(name: str, size_label: str = "") -> str:
    base = f"SS {name}"
    if size_label:
        base = f"{base} {size_label}"
    slug = re.sub(r"[^A-Za-z0-9]+", "-", base.upper()).strip("-")
    return re.sub(r"-{2,}", "-", slug)[:90]


def should_skip_name(name: str) -> bool:
    n = clean(name)
    if len(n) < 2:
        return True
    if SKIP_NAMES.match(n):
        return True
    if re.fullmatch(r"[\d\s./-]+", n):
        return True
    return False


def match_section(text: str):
    t = clean(text).lower()
    # spaced ENGLISH WILLOW
    if re.search(r"e\s*n\s*g\s*l\s*i\s*s\s*h\s*w\s*i\s*l\s*l\s*o\s*w", t):
        return "CC_EW_BATS", "English Willow Cricket Bat"
    for pattern, cat, label in SECTION_MAP:
        if re.search(pattern, t, re.I):
            return cat, label
    return None


def is_section_header(line: str) -> bool:
    s = clean(line)
    if not s or parse_price(s):
        return False
    if extract_pairs(s):
        return False
    # Product-like titles (e.g. PLAYER BAT) are not section headers
    if re.search(r"\b(bat|glove|pad|ball|helmet|cap|shoe|sock|stump|kitbag|guard)\b", s, re.I) and not re.search(
        r"\b(bats|gloves|pads|balls|helmets|caps|shoes|socks|stumps|kitbags|guards|range|series|willow|junior|clothing)\b",
        s,
        re.I,
    ):
        return False
    return match_section(s) is not None or bool(
        re.search(
            r"\b(range|series|bats|gloves|pads|balls|kitbags?|helmets?|caps?|hats?|shoes?|socks?|stumps?|whites?|supporters?|guards?|willow|junior|clothing|sunglasses|grips|tape)\b",
            s,
            re.I,
        )
    )


def extract_pairs(line: str) -> list[tuple[str, int]]:
    line = clean(line)
    if not line:
        return []

    if re.fullmatch(r"(\d{3,6}\s*)+", line):
        return []

    m2 = re.match(r"^(.+?)\s+(\d{3,6})\s+(.+?)\s+(\d{3,6})$", line)
    if m2:
        n1, p1, n2, p2 = clean(m2.group(1)), int(m2.group(2)), clean(m2.group(3)), int(m2.group(4))
        # Model-number lines: "GUNTHER 200000 RESERVE EDITION 95000"
        # Only merge when left token is a short model family name (not a full 2-column product).
        continuation = re.match(
            r"^(reserve(\s+edition)?|player(\s+edition)?|special(\s+edition)?|edition)\b",
            n2,
            re.I,
        )
        if continuation and len(n1.split()) <= 2 and "/" not in n1 and p1 >= 100000:
            merged = clean(f"{n1} {p1} {n2}")
            if not should_skip_name(merged):
                return [(merged, p2)]
        out = []
        for name, price in ((n1, p1), (n2, p2)):
            if should_skip_name(name):
                continue
            if re.search(r"\b(RANGE|SERIES)$", name, re.I):
                continue
            out.append((name, price))
        if out:
            return out

    m1 = re.match(r"^(.+)\s+(\d{3,6})$", line)
    if not m1:
        return []
    name, price = clean(m1.group(1)), int(m1.group(2))
    if should_skip_name(name):
        return []
    if re.search(r"\b(RANGE|SERIES)$", name, re.I):
        return []
    if len(re.findall(r"\b\d{3,6}\b", line)) > 2:
        return []
    return [(name, price)]


def matrix_products(table, category_id, type_label, section) -> list[ProductRow]:
    rows: list[ProductRow] = []
    if not table or len(table) < 2:
        return rows
    first = [clean(str(c or "")) for c in table[0]]
    header_join = " ".join(first).upper()
    if not any(k in header_join for k in ("SIZE", "MEN", "YOUTH", "BOYS", "HARROW", "FULL", "NO.")):
        # simple 2-col table
        for raw in table:
            cells = [clean(str(c or "")) if c is not None else "" for c in raw]
            if len(cells) < 2 or not cells[0]:
                continue
            if should_skip_name(cells[0]):
                continue
            price = parse_price(cells[1])
            if price is None:
                continue
            # Avoid model-number-as-name with absurd split: name looks like edition fragment
            rows.append(
                ProductRow(
                    name=cells[0],
                    mrp=price,
                    category_id=category_id,
                    product_type_label=type_label,
                    section=section,
                    source="table",
                )
            )
        return rows

    size_headers = first[1:]
    for raw in table[1:]:
        cells = [clean(str(c or "")) if c is not None else "" for c in raw]
        if not cells or not cells[0] or should_skip_name(cells[0]):
            continue
        name = cells[0]
        for idx, cell in enumerate(cells[1:]):
            price = parse_price(cell)
            if price is None:
                continue
            size = size_headers[idx] if idx < len(size_headers) else f"Opt-{idx+1}"
            if not size or size in {"-", "—"}:
                continue
            rows.append(
                ProductRow(
                    name=name,
                    mrp=price,
                    category_id=category_id,
                    product_type_label=type_label,
                    size_label=size,
                    section=section,
                    source="matrix",
                    attrs={"size": size},
                )
            )
    return rows


def extract_products() -> list[ProductRow]:
    products: list[ProductRow] = []

    with pdfplumber.open(PDF_PATH) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            if page_idx == 1:
                continue
            text = page.extract_text() or ""
            if "Terms & Conditions" in text or "Terms and Conditions" in text:
                continue

            cat, label = PAGE_DEFAULTS.get(page_idx, ("CC_EW_BATS", "Cricket Product"))
            section = f"Page {page_idx}"

            # Walk text lines for section changes + products
            for line in text.splitlines():
                s = clean(line)
                if not s or s.isdigit():
                    continue

                matched = match_section(s) if (is_section_header(s) or match_section(s)) and not extract_pairs(s) else None
                if matched and (is_section_header(s) or len(s.split()) <= 6):
                    # Only switch on short/header-like lines
                    if not extract_pairs(s):
                        cat, label = matched
                        section = s
                        continue

                # special willow markers even mid-page
                if re.search(r"kashmir\s*willow", s, re.I) and not extract_pairs(s):
                    cat, label = "CC_KW_BATS", "Kashmir Willow Cricket Bat"
                    section = "Kashmir Willow"
                    continue
                if re.search(r"e\s*n\s*g\s*l\s*i\s*s\s*h\s*w\s*i\s*l\s*l\s*o\s*w|english\s*willow", s, re.I) and not extract_pairs(s):
                    cat, label = "CC_EW_BATS", "English Willow Cricket Bat"
                    section = "English Willow"
                    continue

                for name, price in extract_pairs(s):
                    products.append(
                        ProductRow(
                            name=name,
                            mrp=price,
                            category_id=cat,
                            product_type_label=label,
                            section=section,
                            source="text",
                        )
                    )

            # Matrix tables only (size grids) — simple 2-col tables often corrupt multi-column layouts
            for table in page.extract_tables() or []:
                if not table:
                    continue
                width = max(len(r) for r in table)
                if width >= 3:
                    products.extend(matrix_products(table, cat, label, section))
                elif page_idx >= 7:
                    # accessory pages: 2-col tables are usually reliable
                    products.extend(matrix_products(table, cat, label, section))

    return products


def infer_category_from_name(name: str, fallback_cat: str, fallback_label: str):
    n = name.lower()
    rules = [
        (r"helmet|grill", "CC_HELMET", "Cricket Helmet"),
        (r"\bcap\b|super cap", "CC_CAPS", "Cricket Cap"),
        (r"panama|bucket hat|\bhat\b", "CC_HATS", "Cricket Hat"),
        (r"wheelie", "CC_WHEELIE_KITBAGS", "Wheelie Cricket Kitbag"),
        (r"duffle|kitbag|kit bag", "CC_DUFFLE_KITBAGS", "Cricket Kitbag"),
        (r"batting glove", "CC_BATTING_GLOVES", "Cricket Batting Gloves"),
        (r"keeping glove|wicket keep", "CC_KEEPING_GLOVES", "Wicket Keeping Gloves"),
        (r"inner glove|skull cap", "CC_INNER_GLOVES", "Inner Cricket Gloves"),
        (r"batting pad|keeping pad", "CC_BATTING_PADS", "Cricket Pads"),
        (r"thigh", "CC_THIGH_PADS", "Cricket Thigh Pad"),
        (r"elbow", "CC_ELBOW_GUARDS", "Cricket Elbow Guard"),
        (r"supporter|abdo", "CC_SUPPORTERS", "Cricket Supporter"),
        (r"\bstump\b|mallet|rubber base", "CC_STUMPS", "Cricket Stumps"),
        (r"\bshoe\b|spikes", "CC_SHOES", "Cricket Shoes"),
        (r"\bsock", "CC_SOCKS", "Cricket Socks"),
        (r"combo white", "CC_COMBO_WHITES", "Cricket Combo Whites"),
        (r"cricket white|\bwhites\b", "CC_WHITES", "Cricket Whites"),
        (r"leather ball", "CC_LEATHER_BALLS", "Leather Cricket Ball"),
        (r"tennis ball", "CC_TENNIS_BALLS", "Tennis Cricket Ball"),
        (r"synthetic ball", "CC_SYNTHETIC_BALLS", "Synthetic Cricket Ball"),
        (r"practice ball", "CC_PRACTICE_BALLS", "Practice Cricket Ball"),
        (r"plastic|tennis ball bat|soft pro|scoop|composite", "CC_TENNIS_BATS", "Tennis Ball Cricket Bat"),
        (r"kashmir|\bkw\b", "CC_KW_BATS", "Kashmir Willow Cricket Bat"),
        (r"english|\bew\b|willow|bat", "CC_EW_BATS", "English Willow Cricket Bat"),
    ]
    for pat, cat, label in rules:
        if re.search(pat, n, re.I):
            return cat, label
    return fallback_cat, fallback_label


def dedupe_products(products: list[ProductRow]) -> list[ProductRow]:
    # Prefer text-sourced longer names over short table fragments
    by_id: dict[str, ProductRow] = {}

    def score(p: ProductRow) -> tuple:
        return (
            1 if p.source == "text" else 0,
            1 if p.source == "matrix" else 0,
            len(p.name),
            1 if p.size_label else 0,
        )

    for p in products:
        pretty = title_case(p.name)
        if should_skip_name(pretty):
            continue
        if p.mrp < 50 or p.mrp > 200000:
            continue
        # model-number mistaken as MRP (e.g. GUNTHER | 200000)
        if p.mrp >= 100000 and len(p.name.split()) <= 2 and p.source == "table":
            continue
        cat, label = infer_category_from_name(pretty, p.category_id, p.product_type_label)
        pid = slug_id(pretty, p.size_label)
        key = f"{pid}::{p.mrp}"
        cur = by_id.get(key)
        candidate = ProductRow(
            name=pretty,
            mrp=p.mrp,
            category_id=cat,
            product_type_label=label,
            size_label=p.size_label,
            section=p.section,
            source=p.source,
            attrs=dict(p.attrs),
        )
        if cur is None or score(candidate) > score(cur):
            by_id[key] = candidate

    items = list(by_id.values())

    names = list(items)
    drop = set()
    for p in names:
        pid = id(p)
        if p.size_label:
            continue
        for other in names:
            if p is other or other.size_label:
                continue
            if len(other.name) <= len(p.name):
                continue
            if other.name.lower().endswith(" " + p.name.lower()) and abs(other.mrp - p.mrp) <= max(500, int(p.mrp * 0.1)):
                drop.add(pid)
    items = [p for p in items if id(p) not in drop]

    long_items = [(p.name.lower(), p.mrp, p) for p in items]
    filtered = []
    for p in items:
        n = p.name.lower()
        if p.size_label:
            filtered.append(p)
            continue
        drop_short = False
        for ln, lmrp, other in long_items:
            if n == ln or len(ln) <= len(n):
                continue
            if ln.endswith(" " + n) and abs(lmrp - p.mrp) <= max(500, int(p.mrp * 0.1)):
                drop_short = True
                break
            if ln.startswith(n + " ") and abs(lmrp - p.mrp) <= max(500, int(p.mrp * 0.1)):
                drop_short = True
                break
        if not drop_short:
            filtered.append(p)

    # Disambiguate same name + different MRP by appending MRP to id/name for the lower-priced row
    by_base: dict[str, list[ProductRow]] = {}
    for p in filtered:
        by_base.setdefault(slug_id(p.name, p.size_label), []).append(p)

    unique: list[ProductRow] = []
    for base_id, group in by_base.items():
        group = sorted(group, key=lambda x: -x.mrp)
        unique.append(group[0])
        for extra in group[1:]:
            # Keep distinct sellable SKUs for junior/size variants with different MRP
            suffix = extra.size_label or str(extra.mrp)
            if extra.size_label:
                new_name = extra.name  # already distinct via size in slug
            else:
                new_name = f"{extra.name} ({suffix})"
            unique.append(
                ProductRow(
                    name=new_name,
                    mrp=extra.mrp,
                    category_id=extra.category_id,
                    product_type_label=extra.product_type_label,
                    size_label=extra.size_label,
                    section=extra.section,
                    source=extra.source,
                    attrs=dict(extra.attrs),
                )
            )

    unique.sort(key=lambda x: (x.category_id, x.name, x.size_label))
    return unique


def seo_fields(p: ProductRow) -> dict:
    display = f"{p.name} ({p.size_label})" if p.size_label else p.name
    type_label = p.product_type_label
    product_name = f"SS {display} {type_label} | Buy Online India"
    if len(product_name) > 120:
        product_name = f"SS {display} | {type_label}"
    internal = f"SS {display}"
    description = (
        f"Buy genuine SS {display} {type_label.lower()} online. "
        f"Premium Sareen Sports (SS) cricket gear with MRP ₹{p.mrp:,}. "
        f"Trusted quality for practice and match play in India."
    )
    long_description = (
        f"The SS {display} is an authentic {type_label.lower()} from Sareen Sports Industries (SS). "
        f"Designed for performance and durability, it is a popular choice among club and competitive players. "
        f"Catalog MRP: ₹{p.mrp:,}. Shop SS cricket equipment for reliable on-field performance."
    )
    keywords = ", ".join([
        f"SS {display}",
        f"SS {p.name}",
        type_label.lower(),
        "SS cricket",
        "buy SS cricket online",
        "SS cricket equipment India",
        f"SS {display} price",
        "Sareen Sports",
    ])
    return {
        "internal_name": internal,
        "product_name": product_name,
        "description": description,
        "long_description": long_description,
        "keywords": keywords,
    }


def load_valid_categories() -> set[str]:
    wb = openpyxl.load_workbook(CATEGORY_PATH, data_only=True)
    ws = wb["Categories"]
    return {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}


def write_excel(products: list[ProductRow]) -> None:
    valid = load_valid_categories()
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Products"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, p in enumerate(products, start=2):
        seo = seo_fields(p)
        product_id = slug_id(p.name, p.size_label)
        sku = product_id.replace("SS-", "SKU-SS-", 1)
        category_id = p.category_id if p.category_id in valid else "CC_EW_BATS"
        values = {
            "product_id": product_id,
            "sku": sku,
            "product_type_id": "FINISHED_GOOD",
            "status_id": "ACTIVE",
            "category_ids": category_id,
            "internal_name": seo["internal_name"],
            "brand_name": "SS",
            "product_name": seo["product_name"],
            "description": seo["description"],
            "long_description": seo["long_description"],
            "comments": "Imported from SS_CATALOG_MRP.pdf",
            "keywords": seo["keywords"],
            "is_virtual": "N",
            "is_variant": "N",
            "returnable": "Y",
            "taxable": "Y",
            "charge_shipping": "Y",
            "require_inventory": "Y",
            "introduction_date": "2023-01-01",
            "currency": "INR",
            "tax_rate": 18,
            "LIST_PRICE": p.mrp,
            "MAXIMUM_PRICE": p.mrp,
            "DEFAULT_PRICE": p.mrp,
        }
        if p.size_label:
            values["attr_1_name"] = "size"
            values["attr_1_value"] = p.size_label
        for key, val in values.items():
            if key in col:
                ws.cell(row=idx, column=col[key], value=val)

    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(products)} products -> {OUTPUT_PATH.name}")


def main():
    products = extract_products()
    print("raw", len(products))
    products = dedupe_products(products)
    print("final", len(products))
    from collections import Counter
    print("cats", Counter(p.category_id for p in products))
    for p in products:
        if "Gunther" in p.name or "Player Bat" in p.name or p.name == "Gold Edition":
            print("sample", slug_id(p.name, p.size_label), p.mrp, p.category_id, p.name)
    write_excel(products)


if __name__ == "__main__":
    main()
