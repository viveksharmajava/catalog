"""
Build SS bats product import Excel from ss_catalog.txt.

Output: ss-bats-product-import.xlsx
"""

from __future__ import annotations

import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "ss_catalog.txt"
TEMPLATE = ROOT / "ss-product-import.xlsx"
OUTPUT = ROOT / "ss-bats-product-import.xlsx"

CATEGORY_RE = re.compile(
    r"^Category_ID\s*:\s*(?P<id>[A-Za-z0-9_-]+)\s+Category\s+Name\s*:\s*(?P<name>.+)$",
    re.IGNORECASE,
)
PRICE_RE = re.compile(r"^(?P<name>.+?)\s*(?P<price>\d{3,6})\s*$")

DISPLAY_PREFIX = {
    "CC_EW_BATS": "SS English Willow",
    "CC_KW_BATS": "SS Kashmir Willow",
}


def clean(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip(" -|/"))


def title_case(name: str) -> str:
    name = clean(name)
    keep_upper = {"SS", "EW", "KW", "GG", "VA", "DK", "MS", "JR", "XL", "PRO", "R-7"}
    words = []
    for w in name.split(" "):
        up = w.upper()
        if up in keep_upper:
            words.append(up)
        elif re.fullmatch(r"\d+(\.\d+)?", w) or re.fullmatch(r"\d+/\d+", w):
            words.append(w)
        elif "/" in w:
            words.append("/".join(title_case(p) if p else "" for p in w.split("/")))
        elif w.startswith("(") and w.endswith(")"):
            words.append(f"({title_case(w[1:-1])})")
        else:
            words.append(w[:1].upper() + w[1:].lower() if w else w)
    return " ".join(words)


def slug_id(display_name: str) -> str:
    """Product id always starts with SS- and fits catalog VARCHAR(64)."""
    slug = re.sub(r"[^A-Za-z0-9]+", "-", display_name.upper()).strip("-")
    slug = re.sub(r"-{2,}", "-", slug)
    if not slug.startswith("SS-"):
        slug = f"SS-{slug}"
    # Compact known long prefixes before hard truncate
    if len(slug) > 64:
        slug = slug.replace("SS-KASHMIR-WILLOW-SS-KW-", "SS-KW-")
        slug = slug.replace("SS-KASHMIR-WILLOW-", "SS-KW-")
        slug = slug.replace("SS-ENGLISH-WILLOW-", "SS-EW-")
        slug = slug.replace("CRICKET-KIT-WITH-HELMET", "KIT-HELMET")
        slug = slug.replace("CRICKET-KIT", "KIT")
        slug = slug.replace("-SIZE-", "-SZ-")
    return slug[:64].rstrip("-")



def discounted_price(maximum_price: int) -> int:
    """DEFAULT/LIST = 20% less than MAXIMUM (rounded to nearest rupee)."""
    value = Decimal(maximum_price) * Decimal("0.80")
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def parse_catalog(path: Path):
    products = []
    category_id = None
    category_name = None

    for raw in path.read_text(encoding="utf-8").splitlines():
        line = clean(raw)
        if not line:
            continue

        cat = CATEGORY_RE.match(line)
        if cat:
            category_id = cat.group("id").strip()
            category_name = clean(cat.group("name"))
            continue

        if category_id is None:
            continue

        m = PRICE_RE.match(line)
        if not m:
            continue

        base_name = title_case(m.group("name"))
        price = int(m.group("price"))
        if price < 50:
            continue

        prefix = DISPLAY_PREFIX.get(category_id, "SS")
        display_name = f"{prefix} {base_name}"

        products.append(
            {
                "category_id": category_id,
                "category_name": category_name or category_id,
                "base_name": base_name,
                "display_name": display_name,
                "maximum_price": price,
                "sale_price": discounted_price(price),
            }
        )

    seen = set()
    unique = []
    for p in products:
        pid = slug_id(p["display_name"])
        if pid in seen:
            continue
        seen.add(pid)
        p["product_id"] = pid
        unique.append(p)
    return unique


def seo_fields(p: dict) -> dict:
    display = p["display_name"]
    category_name = p["category_name"]
    willow_type = "English willow" if p["category_id"] == "CC_EW_BATS" else (
        "Kashmir willow" if p["category_id"] == "CC_KW_BATS" else "cricket"
    )

    # No "| Buy Online India"
    product_name = display[:100]
    internal = display[:255]

    description = (
        f"Shop {display} cricket bat online. Genuine Sareen Sports (SS) {willow_type} bat "
        f"with MRP ₹{p['maximum_price']:,}. Built for performance in practice and match play."
    )
    long_description = (
        f"{display} is a premium {willow_type} cricket bat from Sareen Sports Industries (SS). "
        f"Designed for balance, power and durability, it suits club and competitive players. "
        f"Catalog maximum price ₹{p['maximum_price']:,}. Explore SS cricket bats for trusted on-field quality."
    )
    keywords = ", ".join(
        [
            display,
            f"buy {display} online",
            f"SS {p['base_name']}",
            f"{willow_type} cricket bat",
            f"SS {willow_type} bat",
            category_name.lower(),
            "SS cricket bat",
            "Sareen Sports bat",
            f"{display} price",
            "cricket bat online India",
        ]
    )
    return {
        "internal_name": internal,
        "product_name": product_name,
        "description": description,
        "long_description": long_description,
        "keywords": keywords,
    }


def write_excel(products: list[dict]) -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Products"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, p in enumerate(products, start=2):
        seo = seo_fields(p)
        product_id = p["product_id"]
        sku = product_id.replace("SS-", "SKU-SS-", 1)
        values = {
            "product_id": product_id,
            "sku": sku,
            "product_type_id": "FINISHED_GOOD",
            "status_id": "ACTIVE",
            "category_ids": p["category_id"],
            "internal_name": seo["internal_name"],
            "brand_name": "SS",
            "product_name": seo["product_name"],
            "description": seo["description"],
            "long_description": seo["long_description"],
            "comments": "Imported from ss_catalog.txt",
            "keywords": seo["keywords"],
            "is_virtual": "N",
            "is_variant": "N",
            "returnable": "Y",
            "taxable": "Y",
            "charge_shipping": "Y",
            "require_inventory": "Y",
            "introduction_date": "2023-01-01",
            "small_image": f"{product_id}/small/main.jpg",
            "medium_image": f"{product_id}/medium/main.jpg",
            "large_image": f"{product_id}/large/main.jpg",
            "currency": "INR",
            "tax_rate": 18,
            "DEFAULT_PRICE": p["sale_price"],
            "LIST_PRICE": p["sale_price"],
            "MAXIMUM_PRICE": p["maximum_price"],
        }
        for key, val in values.items():
            if key in col:
                ws.cell(row=idx, column=col[key], value=val)

    wb.save(OUTPUT)


def main():
    products = parse_catalog(SOURCE)
    write_excel(products)
    print(f"Parsed {len(products)} products from {SOURCE.name}")
    print(f"Wrote {OUTPUT.name}")
    for p in products[:3]:
        print(
            f"  {p['product_id']} | max={p['maximum_price']} "
            f"default/list={p['sale_price']} | {p['display_name']}"
        )
    kw = [p for p in products if p["category_id"] == "CC_KW_BATS"][:2]
    for p in kw:
        print(
            f"  {p['product_id']} | max={p['maximum_price']} "
            f"default/list={p['sale_price']} | {p['display_name']}"
        )


if __name__ == "__main__":
    main()
